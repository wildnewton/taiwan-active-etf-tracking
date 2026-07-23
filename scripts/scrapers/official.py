"""Official ETF issuer fallback scrapers.

Official sources share the same structural snapshot validation as MoneyDJ.
Total weight remains a diagnostic warning and never determines validity.
"""

import asyncio
import json
import re
from datetime import date, datetime
from io import BytesIO
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from playwright.async_api import Error as PlaywrightError
from playwright.async_api import TimeoutError as PlaywrightTimeoutError

from config import get_etf_config
from scrapers.moneydj import classify_asset, dedupe_rows, split_rows
from snapshot_validation import validate_snapshot_rows


SOURCE_TYPE = "official_fallback"
OFFICIAL_WARNING_MIN_TOTAL_WEIGHT = 20.0
OFFICIAL_WARNING_MAX_TOTAL_WEIGHT = 110.0
EXTRACTION_METHOD_STATIC = "requests_bs4"
EXTRACTION_METHOD_API = "playwright_api_intercept"
EXTRACTION_METHOD_EXCEL = "requests_xlsx"
EXTRACTION_METHOD_PLAYWRIGHT = "playwright_table_parse"
EXTRACTION_METHOD_STEALTH = "stealth_playwright_api"
_API_RESPONSE_TIMEOUT_MS = 10_000
_ALLIANZ_FUND_OPTIONS_PATH = "/webapi/api/Category/GetFundDropdownOptions"
_ALLIANZ_TRADE_INFO_PATH = "/webapi/api/Fund/GetFundTradeInfo"
_ALLIANZ_COMBOBOX_SELECTOR = '[role="combobox"][aria-label*="主動安聯"]'

TWSE_URL_TEMPLATE = (
    "https://www.twse.com.tw/zh/products/securities/etf/products/content.html?{code}="
)


def get_official_config(etf_code: str) -> dict:
    etf = get_etf_config(etf_code.upper())
    internal_ids = _parse_official_logic(etf.get("official_logic", ""))
    return {
        "code": etf["code"],
        "issuer": etf["issuer"],
        "name": etf["name"],
        "url": etf["official_url"],
        "method": etf["official_method"],
        "official_logic": etf.get("official_logic"),
        "internal_id": next(iter(internal_ids.values()), None),
        "internal_ids": internal_ids,
    }


def fetch_static(url: str, timeout: int = 30) -> str:
    parsed = urlparse(url)
    referer = f"{parsed.scheme}://{parsed.netloc}/" if parsed.scheme and parsed.netloc else url
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
        ),
        "Accept": (
            "text/html,application/xhtml+xml,application/xml;q=0.9,"
            "image/avif,image/webp,*/*;q=0.8"
        ),
        "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7",
        "Referer": referer,
    }
    response = requests.get(url, headers=headers, timeout=timeout)
    response.raise_for_status()
    response.encoding = response.apparent_encoding or "utf-8"
    return response.text


# Static parsers

def parse_fubon(html: str, etf_code: str, source_url: str) -> list[dict]:
    return _parse_official_table(html, etf_code, source_url)


def parse_taishin(html: str, etf_code: str, source_url: str) -> list[dict]:
    return _parse_official_table(html, etf_code, source_url)


def parse_twse(html: str, etf_code: str, source_url: str) -> list[dict]:
    return _parse_official_table(html, etf_code, source_url)


# API / text parsers

def parse_capital_api(buyback_json: str, etf_code: str, source_url: str) -> list[dict]:
    """Parse Capital's buyback API response.

    The observed payload uses data.stocks[] with keys such as stocNo, stocName,
    share, weight, and weightRound. Keep fallbacks for nearby naming variants.
    """
    data = json.loads(buyback_json)
    payload = data.get("data", {}) if isinstance(data, dict) else {}
    stocks = payload.get("stocks", []) if isinstance(payload, dict) else []
    date = None
    pcf = payload.get("pcf", {}) if isinstance(payload, dict) else {}
    if isinstance(pcf, dict):
        date = pcf.get("date2") or pcf.get("date1")
    date = _normalize_date(date)

    rows = []
    for item in stocks:
        if not isinstance(item, dict):
            continue
        code = str(
            item.get("stocNo")
            or item.get("stockNo")
            or item.get("stockCode")
            or item.get("code")
            or ""
        ).strip()
        name = str(
            item.get("stocName")
            or item.get("stockName")
            or item.get("name")
            or ""
        ).strip()
        shares = _parse_number(str(item.get("share") or item.get("shares") or item.get("qty") or ""))
        weight = _parse_float(str(item.get("weightRound") or item.get("weight") or item.get("ratio") or ""))
        code_match = re.search(r"\b(\d{4})\b", code)
        if not code_match or not name or weight is None:
            continue
        rows.append(_row(etf_code, code_match.group(1), name, shares, weight, source_url, date, EXTRACTION_METHOD_API))
    return rows


def parse_nomura_api(assets_json: str, etf_code: str, source_url: str) -> list[dict]:
    data = json.loads(assets_json)
    entries = data.get("Entries", {}) if isinstance(data, dict) else {}
    fund_id = entries.get("FundID") if isinstance(entries, dict) else None
    if fund_id and str(fund_id).strip().upper() != etf_code.upper():
        raise ValueError(
            f"Nomura fund mismatch: expected {etf_code.upper()}, got {fund_id}"
        )
    fund_data = entries.get("Data", {}) if isinstance(entries, dict) else {}
    nav_date = fund_data.get("FundAsset", {}).get("NavDate")
    date = _normalize_date(nav_date)

    rows = []
    for table in fund_data.get("Table", []):
        if table.get("TableTitle") != "股票":
            continue
        for raw in table.get("Rows", []):
            if len(raw) < 4:
                continue
            code, name, shares, weight = raw[:4]
            code_match = re.search(r"\b(\d{4})\b", str(code))
            parsed_weight = _parse_float(str(weight))
            if not code_match or not name or parsed_weight is None:
                continue
            rows.append(
                _row(
                    etf_code,
                    code_match.group(1),
                    str(name).strip(),
                    _parse_number(str(shares)),
                    parsed_weight,
                    source_url,
                    date,
                    EXTRACTION_METHOD_STEALTH,
                )
            )
    return rows


def parse_ctbc_api(api_json: str, etf_code: str, source_url: str) -> list[dict]:
    data = json.loads(api_json)
    payload = data.get("Data", {}) if isinstance(data, dict) else {}
    assets = payload.get("FundAssets", []) if isinstance(payload, dict) else []
    date = None
    if assets and isinstance(assets[0], dict):
        date = assets[0].get("資料日期")
    date = _normalize_date(date)

    rows = []
    details = payload.get("FundAssetsDetail", []) if isinstance(payload, dict) else []
    for group in details:
        if not isinstance(group, dict) or str(group.get("Code") or "").upper() != "STOCK":
            continue
        for item in group.get("Data", []):
            if not isinstance(item, dict):
                continue
            code = str(item.get("code_") or "").strip()
            name = str(item.get("name_") or "").strip()
            shares = _parse_float(str(item.get("qty_") or ""))
            weight = _parse_float(str(item.get("weights_") or ""))
            code_match = re.search(r"\b(\d{4})\b", code)
            if not code_match or not name or weight is None:
                continue
            rows.append(
                _row(
                    etf_code,
                    code_match.group(1),
                    name,
                    shares,
                    weight,
                    source_url,
                    date,
                    EXTRACTION_METHOD_API,
                )
            )
    return dedupe_rows(rows)


def parse_allianz_fund_options(options_json: str, etf_code: str) -> str:
    """Return Allianz's internal FundNo for one exact exchange ETF code."""
    data = json.loads(options_json)
    if not isinstance(data, dict) or data.get("StatusCode") != 0:
        message = data.get("Message") if isinstance(data, dict) else "invalid payload"
        raise ValueError(f"Allianz fund options API failed: {message}")
    entries = data.get("Entries", [])
    requested_code = etf_code.upper()
    matches = [
        item
        for item in entries
        if isinstance(item, dict)
        and str(item.get("SecuritiesCode") or "").strip().upper() == requested_code
    ]
    if len(matches) != 1:
        raise ValueError(
            f"Allianz fund option not found or ambiguous for {requested_code}"
        )
    fund_no = str(matches[0].get("FundNo") or "").strip()
    if not fund_no:
        raise ValueError(f"Allianz FundNo missing for {requested_code}")
    return fund_no


def _parse_allianz_trade_entries(trade_json: str) -> dict:
    data = json.loads(trade_json)
    if not isinstance(data, dict) or data.get("StatusCode") != 0:
        message = data.get("Message") if isinstance(data, dict) else "invalid payload"
        raise ValueError(f"Allianz trade info API failed: {message}")
    entries = data.get("Entries", {})
    if not isinstance(entries, dict):
        raise ValueError("Allianz trade response entries missing")
    return entries


def _parse_allianz_trade_identity(trade_json: str) -> tuple[str, str]:
    entries = _parse_allianz_trade_entries(trade_json)
    etf_code = str(entries.get("CSecuritiesCode") or "").strip().upper()
    fund_no = str(entries.get("CFundId") or "").strip()
    if not etf_code or not fund_no:
        raise ValueError("Allianz trade response identity missing")
    return etf_code, fund_no


def parse_allianz_api(
    trade_json: str,
    etf_code: str,
    source_url: str,
    *,
    expected_fund_no: str,
) -> list[dict]:
    """Parse one exact Allianz fund trade-info response."""
    entries = _parse_allianz_trade_entries(trade_json)

    requested_code = etf_code.upper()
    response_code = str(entries.get("CSecuritiesCode") or "").strip().upper()
    if response_code != requested_code:
        raise ValueError(
            f"Allianz ETF mismatch: expected {requested_code}, "
            f"got {response_code or 'missing'}"
        )

    response_fund_no = str(entries.get("CFundId") or "").strip()
    if response_fund_no != expected_fund_no:
        raise ValueError(
            f"Allianz fund mismatch: expected {expected_fund_no}, "
            f"got {response_fund_no or 'missing'}"
        )

    raw_date = entries.get("CPcfdate")
    date = _normalize_date(str(raw_date).split("T", 1)[0]) if raw_date else None
    if not date:
        raise ValueError("Allianz holdings date missing")
    try:
        datetime.strptime(date, "%Y/%m/%d")
    except ValueError as exc:
        raise ValueError(f"Allianz holdings date invalid: {raw_date}") from exc

    tables = entries.get("DynamicTableData")
    if not isinstance(tables, list):
        raise ValueError("Allianz holdings tables missing")

    stock_table = None
    for table in tables:
        if not isinstance(table, dict):
            continue
        title = str(table.get("TableTitle") or "").strip()
        if title.startswith("股票"):
            stock_table = table
            break
    if stock_table is None:
        raise ValueError("Allianz stock table not found")

    columns = stock_table.get("Columns")
    if not isinstance(columns, list):
        raise ValueError("Allianz stock table schema invalid")
    headers = [
        _normalize_header(str(column.get("Name") or ""))
        for column in columns
        if isinstance(column, dict)
    ]
    header_map = _build_header_map(headers)
    required_fields = {"code", "name", "shares", "weight"}
    if not required_fields.issubset(header_map):
        raise ValueError("Allianz stock table schema invalid")

    raw_rows = stock_table.get("Rows")
    if not isinstance(raw_rows, list):
        raise ValueError("Allianz stock table rows invalid")

    rows = []
    for raw in raw_rows:
        if not isinstance(raw, list):
            continue
        values = _extract_cells([str(value) for value in raw], header_map)
        if not values:
            continue
        stock_code, stock_name, shares, weight_pct = values
        rows.append(
            _row(
                requested_code,
                stock_code,
                stock_name,
                shares,
                weight_pct,
                source_url,
                date,
                EXTRACTION_METHOD_API,
            )
        )
    if not rows:
        raise ValueError("Allianz stock rows not found")
    return dedupe_rows(rows)

def parse_mega_text(body_text: str, etf_code: str, source_url: str, date: str | None = None) -> list[dict]:
    if not date:
        match = re.search(r"(\d{4}/\d{2}/\d{2})", body_text)
        date = match.group(1) if match else None

    rows = []
    lines = [line.strip() for line in body_text.splitlines() if line.strip()]
    for index, line in enumerate(lines):
        if not re.fullmatch(r"\d{4}", line):
            continue
        try:
            code = line
            name = lines[index + 1]
            shares = _parse_number(lines[index + 2])
            weight = _parse_float(lines[index + 3])
        except IndexError:
            continue
        if not name or weight is None:
            continue
        rows.append(_row(etf_code, code, name, shares, weight, source_url, date, EXTRACTION_METHOD_PLAYWRIGHT))
    return rows


def parse_uni_president_table(
    table_rows: list[list[str]],
    etf_code: str,
    source_url: str,
    date: str | None = None,
) -> list[dict]:
    rows = []
    for cells in table_rows:
        if len(cells) < 4:
            continue
        code, name, shares, weight = cells[:4]
        code_match = re.search(r"\b(\d{4})\b", str(code))
        parsed_weight = _parse_float(str(weight))
        if not code_match or not name or parsed_weight is None:
            continue
        rows.append(
            _row(
                etf_code,
                code_match.group(1),
                str(name).strip(),
                _parse_number(str(shares)),
                parsed_weight,
                source_url,
                date,
                EXTRACTION_METHOD_PLAYWRIGHT,
            )
        )
    return rows


# Browser / API scraper functions

async def scrape_capital_playwright(etf_code: str, page) -> dict:
    etf_code = etf_code.upper()
    config = get_official_config(etf_code)
    source_url = config["url"]
    navigation_completed = False

    try:
        async with page.expect_response(
            _is_capital_buyback_response,
            timeout=_API_RESPONSE_TIMEOUT_MS,
        ) as response_info:
            await page.goto(source_url, wait_until="domcontentloaded", timeout=60000)
            navigation_completed = True
        response = await response_info.value
    except PlaywrightTimeoutError:
        if not navigation_completed:
            raise
        return _failed_result(source_url, "Capital buyback API not intercepted")
    except PlaywrightError:
        if not navigation_completed:
            raise
        return _failed_result(source_url, "Capital buyback API not intercepted")

    try:
        buyback_body = await response.text()
    except Exception:
        return _failed_result(source_url, "Capital buyback API not intercepted")

    try:
        all_rows = dedupe_rows(parse_capital_api(buyback_body, etf_code, source_url))
    except Exception as exc:
        return _failed_result(source_url, f"Capital API parse error: {exc}")

    return _build_result(all_rows, source_url, EXTRACTION_METHOD_API)


async def scrape_nomura_stealth(etf_code: str, page) -> dict:
    etf_code = etf_code.upper()
    config = get_official_config(etf_code)
    source_url = config["url"]
    navigation_completed = False

    try:
        async with page.expect_response(
            _is_nomura_assets_response,
            timeout=_API_RESPONSE_TIMEOUT_MS,
        ) as response_info:
            await page.goto(source_url, wait_until="domcontentloaded", timeout=60000)
            navigation_completed = True
        response = await response_info.value
    except PlaywrightTimeoutError:
        if not navigation_completed:
            raise
        return _failed_result(source_url, "Nomura GetFundAssets API not intercepted")
    except PlaywrightError:
        if not navigation_completed:
            raise
        return _failed_result(source_url, "Nomura GetFundAssets API not intercepted")

    try:
        assets_body = await response.text()
    except Exception:
        return _failed_result(source_url, "Nomura GetFundAssets API not intercepted")

    try:
        all_rows = dedupe_rows(parse_nomura_api(assets_body, etf_code, source_url))
    except Exception as exc:
        return _failed_result(source_url, f"Nomura API parse error: {exc}")
    return _build_result(all_rows, source_url, EXTRACTION_METHOD_STEALTH)


async def scrape_ctbc_playwright(etf_code: str, page) -> dict:
    etf_code = etf_code.upper()
    config = get_official_config(etf_code)
    source_url = config["url"]
    navigation_completed = False

    try:
        async with page.expect_response(
            _is_ctbc_holdings_response,
            timeout=_API_RESPONSE_TIMEOUT_MS,
        ) as response_info:
            await page.goto(source_url, wait_until="domcontentloaded", timeout=60000)
            navigation_completed = True
        response = await response_info.value
    except PlaywrightTimeoutError:
        if not navigation_completed:
            raise
        return _failed_result(source_url, "CTBC ETFHoldingWeight API not intercepted")
    except PlaywrightError:
        if not navigation_completed:
            raise
        return _failed_result(source_url, "CTBC ETFHoldingWeight API not intercepted")

    try:
        holdings_body = await response.text()
    except Exception:
        return _failed_result(source_url, "CTBC ETFHoldingWeight API not intercepted")

    try:
        all_rows = dedupe_rows(parse_ctbc_api(holdings_body, etf_code, source_url))
    except Exception as exc:
        return _failed_result(source_url, f"CTBC API parse error: {exc}")

    return _build_result(all_rows, source_url, EXTRACTION_METHOD_API)


async def _allianz_response_text(response, label: str) -> str:
    if getattr(response, "ok", True) is False:
        raise ValueError(
            f"Allianz {label} API HTTP {getattr(response, 'status', 'error')}"
        )
    return await response.text()


async def _switch_allianz_fund(etf_code: str, page) -> str:
    combobox = page.locator(_ALLIANZ_COMBOBOX_SELECTOR)
    if await combobox.count() != 1:
        raise ValueError(f"Allianz fund selector not found for {etf_code}")
    await combobox.click()

    option_selector = f'[role="option"][aria-label^="{etf_code} "]'
    option = page.locator(option_selector)
    try:
        await option.wait_for(state="visible", timeout=_API_RESPONSE_TIMEOUT_MS)
    except (PlaywrightTimeoutError, PlaywrightError) as exc:
        raise ValueError(f"Allianz fund option not found for {etf_code}") from exc
    if await option.count() != 1:
        raise ValueError(
            f"Allianz fund option not found or ambiguous for {etf_code}"
        )

    async with page.expect_response(
        _is_allianz_trade_info_response,
        timeout=_API_RESPONSE_TIMEOUT_MS,
    ) as response_info:
        await option.click()
    response = await response_info.value
    return await _allianz_response_text(response, "trade info")


async def scrape_allianz_playwright(etf_code: str, page) -> dict:
    etf_code = etf_code.upper()
    config = get_official_config(etf_code)
    source_url = config["url"]
    navigation_completed = False

    try:
        async with page.expect_response(
            _is_allianz_fund_options_response,
            timeout=_API_RESPONSE_TIMEOUT_MS,
        ) as options_info:
            async with page.expect_response(
                _is_allianz_trade_info_response,
                timeout=_API_RESPONSE_TIMEOUT_MS,
            ) as trade_info:
                await page.goto(
                    source_url,
                    wait_until="domcontentloaded",
                    timeout=60000,
                )
                navigation_completed = True
        options_response = await options_info.value
        initial_trade_response = await trade_info.value
    except (PlaywrightTimeoutError, PlaywrightError):
        if not navigation_completed:
            raise
        return _failed_result(
            source_url,
            "Allianz initial APIs not intercepted",
        )

    try:
        options_body = await _allianz_response_text(
            options_response,
            "fund options",
        )
        expected_fund_no = parse_allianz_fund_options(options_body, etf_code)

        initial_trade_body = await _allianz_response_text(
            initial_trade_response,
            "trade info",
        )
        initial_code, initial_fund_no = _parse_allianz_trade_identity(
            initial_trade_body
        )
        if (initial_code, initial_fund_no) == (etf_code, expected_fund_no):
            trade_body = initial_trade_body
        else:
            trade_body = await _switch_allianz_fund(etf_code, page)

        all_rows = parse_allianz_api(
            trade_body,
            etf_code,
            source_url,
            expected_fund_no=expected_fund_no,
        )
    except Exception as exc:
        return _failed_result(source_url, f"Allianz API error: {exc}")

    return _build_result(all_rows, source_url, EXTRACTION_METHOD_API)


async def scrape_mega_playwright(etf_code: str, page) -> dict:
    etf_code = etf_code.upper()
    config = get_official_config(etf_code)
    source_url = config["url"]

    await page.goto(source_url, wait_until="load", timeout=60000)
    await page.wait_for_timeout(3000)
    body_text = await page.locator("body").inner_text()

    all_rows = dedupe_rows(parse_mega_text(body_text, etf_code, source_url))
    return _build_result(all_rows, source_url, EXTRACTION_METHOD_PLAYWRIGHT)


async def scrape_uni_president_playwright(etf_code: str, page) -> dict:
    etf_code = etf_code.upper()
    config = get_official_config(etf_code)
    source_url = config["url"]

    await page.goto(source_url, wait_until="load", timeout=60000)
    try:
        await page.wait_for_selector("table", timeout=10000)
    except Exception:
        return _failed_result(source_url, "Uni-President holdings table not found")

    tables = await page.query_selector_all("table")
    table_data = []
    date = None
    for table in tables:
        rows = await table.query_selector_all("tr")
        if len(rows) < 20:
            continue
        first_row_text = await rows[0].inner_text()
        if "股票" not in first_row_text:
            continue
        for row in rows[1:]:
            cells = await row.query_selector_all("td")
            cell_texts = [(await cell.inner_text()).strip() for cell in cells]
            if len(cell_texts) >= 4:
                table_data.append(cell_texts[:4])
        pane_text = await _uni_president_portfolio_pane_text(table)
        date = _parse_uni_president_holdings_date(pane_text)
        break

    if not table_data:
        return _failed_result(source_url, "Uni-President holdings table not found")
    if not date:
        return _failed_result(
            source_url,
            "Uni-President holdings date not found in portfolio pane",
        )

    all_rows = dedupe_rows(parse_uni_president_table(table_data, etf_code, source_url, date))
    return _build_result(all_rows, source_url, EXTRACTION_METHOD_PLAYWRIGHT)


# Unified entry points

def scrape_official_static(etf_code: str) -> dict:
    etf_code = etf_code.upper()
    source_url = _build_twse_url(etf_code)

    try:
        config = get_official_config(etf_code)
        if config["method"] == "static":
            source_url = config["url"]
            parser = _parser_for_issuer(config["issuer"])
        else:
            parser = parse_twse
        html = fetch_static(source_url)
        all_rows = dedupe_rows(parser(html, etf_code, source_url))
        return _build_result(all_rows, source_url, EXTRACTION_METHOD_STATIC)
    except KeyError:
        try:
            html = fetch_static(source_url)
            all_rows = dedupe_rows(parse_twse(html, etf_code, source_url))
            return _build_result(all_rows, source_url, EXTRACTION_METHOD_STATIC)
        except Exception as exc:
            return _failed_result(source_url, str(exc))
    except Exception as exc:
        return _failed_result(source_url, str(exc))


async def scrape_official_with_browser(
    etf_code: str,
    page,
    target_date: date | None = None,
) -> dict:
    etf_code = etf_code.upper()
    config = get_official_config(etf_code)
    method = config["method"]
    issuer = config["issuer"]

    if method == "api" and issuer == "JPMorgan":
        if target_date is None:
            return _failed_result(config["url"], "target_date is required for JPMorgan")
        return await asyncio.to_thread(scrape_jpmorgan_excel, etf_code, target_date)

    if method == "api" and issuer == "Capital":
        return await scrape_capital_playwright(etf_code, page)
    if method == "stealth_api" and issuer == "Nomura":
        return await scrape_nomura_stealth(etf_code, page)
    if method == "browser" and issuer == "CTBC":
        return await scrape_ctbc_playwright(etf_code, page)
    if method == "playwright" and issuer == "Allianz":
        return await scrape_allianz_playwright(etf_code, page)
    if method == "playwright" and issuer == "Mega":
        return await scrape_mega_playwright(etf_code, page)
    if method == "playwright" and issuer == "Uni-President":
        return await scrape_uni_president_playwright(etf_code, page)


    return _failed_result(config["url"], f"No browser official scraper for {issuer} (method={method})")


# Internal helpers

def _response_url(response) -> str:
    url = getattr(response, "url", "")
    return url if isinstance(url, str) else ""


def _has_expected_response_method(response, expected_method: str) -> bool:
    if getattr(response, "ok", False) is not True:
        return False
    request = getattr(response, "request", None)
    method = getattr(request, "method", None)
    return (
        isinstance(method, str)
        and method.upper() == expected_method.upper()
    )


def _matches_api_endpoint(
    response,
    domain: str,
    path: str,
    expected_method: str,
) -> bool:
    parsed = urlparse(_response_url(response))
    hostname = (parsed.hostname or "").lower()
    response_path = parsed.path.rstrip("/").lower()
    expected_domain = domain.lower()
    host_matches = hostname == expected_domain or hostname.endswith(
        f".{expected_domain}"
    )
    return (
        host_matches
        and response_path == path.lower()
        and _has_expected_response_method(response, expected_method)
    )


def _is_capital_buyback_response(response) -> bool:
    return _matches_api_endpoint(
        response,
        "capitalfund.com.tw",
        "/cfweb/api/etf/buyback",
        "POST",
    )


def _is_nomura_assets_response(response) -> bool:
    return _matches_api_endpoint(
        response,
        "nomurafunds.com.tw",
        "/api/etfapi/api/fund/getfundassets",
        "POST",
    )


def _is_ctbc_holdings_response(response) -> bool:
    return _matches_api_endpoint(
        response,
        "ctbcinvestments.com.tw",
        "/api/etf/etfholdingweight",
        "GET",
    )


def _matches_post_api_endpoint(response, domain: str, path: str) -> bool:
    parsed = urlparse(_response_url(response))
    hostname = (parsed.hostname or "").lower()
    response_path = parsed.path.rstrip("/").lower()
    expected_domain = domain.lower()
    host_matches = hostname == expected_domain or hostname.endswith(
        f".{expected_domain}"
    )
    request = getattr(response, "request", None)
    method = getattr(request, "method", "")
    return (
        host_matches
        and response_path == path.lower()
        and isinstance(method, str)
        and method.upper() == "POST"
    )


def _is_allianz_fund_options_response(response) -> bool:
    return _matches_post_api_endpoint(
        response,
        "etf.allianzgi.com.tw",
        _ALLIANZ_FUND_OPTIONS_PATH,
    )


def _is_allianz_trade_info_response(response) -> bool:
    return _matches_post_api_endpoint(
        response,
        "etf.allianzgi.com.tw",
        _ALLIANZ_TRADE_INFO_PATH,
    )


async def _uni_president_portfolio_pane_text(table) -> str:
    """Return hidden text from the pane that owns the matched holdings table."""
    try:
        text = await table.evaluate(
            """
            (table) => {
                const pane = table.closest('.tab-pane, [role="tabpanel"]');
                return pane ? (pane.textContent || '') : '';
            }
            """
        )
    except Exception:
        return ""
    return text if isinstance(text, str) else ""


def _parse_official_table(html: str, etf_code: str, source_url: str) -> list[dict]:
    soup = BeautifulSoup(html, "lxml")
    date = _parse_date(soup)
    rows = []
    for table in soup.find_all("table"):
        rows.extend(_parse_table_rows(table, etf_code.upper(), source_url, date))
    return rows


def _parse_table_rows(table, etf_code: str, source_url: str, date: str | None) -> list[dict]:
    trs = table.find_all("tr")
    if not trs:
        return []

    header_map = {}
    data_trs = trs
    first_cells = trs[0].find_all(["th", "td"])
    if first_cells and (trs[0].find("th") or _looks_like_header(first_cells)):
        headers = [_normalize_header(cell.get_text(" ", strip=True)) for cell in first_cells]
        header_map = _build_header_map(headers)
        data_trs = trs[1:]

    rows = []
    for tr in data_trs:
        cells = [cell.get_text(" ", strip=True) for cell in tr.find_all("td")]
        if len(cells) < 4:
            continue
        values = _extract_cells(cells, header_map)
        if not values:
            continue
        stock_code, stock_name, shares, weight_pct = values
        rows.append(_row(etf_code, stock_code, stock_name, shares, weight_pct, source_url, date, EXTRACTION_METHOD_STATIC))
    return rows


def _extract_cells(cells: list[str], header_map: dict) -> tuple | None:
    if header_map:
        try:
            code_text = cells[header_map["code"]]
            name_text = cells[header_map["name"]]
            shares_text = cells[header_map["shares"]]
            weight_text = cells[header_map["weight"]]
        except (IndexError, KeyError):
            return None
    else:
        code_text, name_text, shares_text, weight_text = cells[:4]

    code_match = re.search(r"\b(\d{4})\b", code_text)
    stock_name = name_text.strip()
    shares = _parse_number(shares_text)
    weight_pct = _parse_float(weight_text)
    if not code_match or not stock_name or weight_pct is None:
        return None
    return code_match.group(1), stock_name, shares, weight_pct


def _row(etf_code, stock_code, stock_name, shares, weight_pct, source_url, date, method):
    asset_name = f"{stock_name}({stock_code}.TW)"
    classification = classify_asset(asset_name)
    return {
        "date": date,
        "etf_code": etf_code.upper(),
        "asset_name": asset_name,
        "asset_type": classification["asset_type"],
        "stock_code": classification["stock_code"],
        "stock_name": classification["stock_name"],
        "shares": shares,
        "weight_pct": weight_pct,
        "source_url": source_url,
        "source_type": SOURCE_TYPE,
        "extraction_method": method,
    }


def _build_header_map(headers: list[str]) -> dict:
    field_patterns = {
        "code": ("股票代號", "股票代碼", "證券代號", "代號", "code"),
        "name": ("股票名稱", "證券名稱", "名稱", "name"),
        "shares": ("持有股數", "持股數", "庫存股數", "股數", "shares"),
        "weight": ("權重", "投資比例", "佔基金淨資產比例", "比例", "weight", "%"),
    }
    header_map = {}
    for field, patterns in field_patterns.items():
        for index, header in enumerate(headers):
            if any(pattern in header for pattern in patterns):
                header_map[field] = index
                break
    return header_map


def _looks_like_header(cells) -> bool:
    text = " ".join(cell.get_text(" ", strip=True) for cell in cells)
    header_terms = ("股票", "證券", "代號", "名稱", "股數", "權重", "比例", "code", "name")
    return any(term in text.lower() for term in header_terms)


def _normalize_header(value: str) -> str:
    return re.sub(r"\s+", "", value).lower()


def _normalize_date(value):
    if not value:
        return None
    return str(value).replace("-", "/")


def _parse_date(soup: BeautifulSoup) -> str | None:
    text = soup.get_text(" ", strip=True)
    data_date_match = re.search(
        r"(?:資料日期|日期)\s*[:：]?\s*(\d{4}/\d{2}/\d{2})",
        text,
    )
    if data_date_match:
        return data_date_match.group(1)
    date_match = re.search(r"\d{4}/\d{2}/\d{2}", text)
    return date_match.group(0) if date_match else None


def _parse_uni_president_holdings_date(pane_text: str) -> str | None:
    """Extract the labeled holdings date from the matched portfolio pane."""
    labeled_date_match = re.search(
        r"(?:投資組合資料日期|投資組合日期|持股資料日期|股票投資明細資料日期|資料日期)\s*[:：]?\s*"
        r"(\d{4}/\d{2}/\d{2})",
        pane_text,
    )
    return labeled_date_match.group(1) if labeled_date_match else None




_JPMORGAN_SHEETS = {
    "基金資產 - 股票": "stock",
    "基金資產 - 期貨": "futures",
    "基金資產 - 選擇權": "options",
    "現金與約當現金": "cash",
}


def _jpmorgan_sheet_rows(sheet, expected_date: date) -> list[tuple[str, ...]]:
    rows = [
        tuple("" if value is None else str(value).strip() for value in row)
        for row in sheet.iter_rows(values_only=True)
    ]
    title = rows[0][0] if rows and rows[0] else ""
    match = re.search(r"\((\d{4}-\d{2}-\d{2})\)", title)
    actual_date = match.group(1) if match else "missing"
    if actual_date != expected_date.isoformat():
        raise ValueError(
            f"JPMorgan date mismatch: expected {expected_date.isoformat()}, "
            f"got {actual_date}"
        )

    for index, row in enumerate(rows):
        if row and row[0] in {"股票代碼", "商品代碼", "名稱"}:
            return [item for item in rows[index + 1 :] if item and item[0]]
    raise ValueError(f"JPMorgan table header not found: {title}")


def _jpmorgan_non_stock_row(
    raw: tuple[str, ...],
    asset_type: str,
    etf_code: str,
    source_url: str,
    date_str: str,
) -> dict | None:
    if asset_type == "cash":
        if len(raw) < 3:
            return None
        code, name, shares, amount, weight = None, raw[0], None, raw[1], raw[2]
        asset_name = name
    else:
        if len(raw) < 4 or raw[0].upper() in {"", "-", "N/A"}:
            return None
        code, name, shares, amount, weight = raw[0], raw[1], raw[2], None, raw[3]
        asset_name = f"{name}({code})"

    parsed_weight = _parse_float(weight)
    if not name or parsed_weight is None:
        return None
    return {
        "date": date_str,
        "etf_code": etf_code,
        "asset_name": asset_name,
        "asset_type": asset_type,
        "stock_code": code,
        "stock_name": name,
        "shares": _parse_number(shares) if shares else None,
        "market_value": _parse_number(amount) if amount else None,
        "weight_pct": parsed_weight,
        "source_url": source_url,
        "source_type": SOURCE_TYPE,
        "extraction_method": EXTRACTION_METHOD_EXCEL,
    }


def parse_jpmorgan_excel(
    content: bytes,
    etf_code: str,
    source_url: str,
    target_date: date,
) -> list[dict]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    missing = [name for name in _JPMORGAN_SHEETS if name not in workbook.sheetnames]
    if missing:
        raise ValueError(f"JPMorgan sheets missing: {', '.join(missing)}")

    etf_code = etf_code.upper()
    date_str = target_date.strftime("%Y/%m/%d")
    rows = []
    for sheet_name, asset_type in _JPMORGAN_SHEETS.items():
        for raw in _jpmorgan_sheet_rows(workbook[sheet_name], target_date):
            if asset_type == "stock":
                if len(raw) < 5 or not re.fullmatch(r"\d{4}", raw[0]):
                    continue
                weight = _parse_float(raw[4])
                if not raw[1] or weight is None:
                    continue
                row = _row(
                    etf_code,
                    raw[0],
                    raw[1],
                    _parse_number(raw[2]),
                    weight,
                    source_url,
                    date_str,
                    EXTRACTION_METHOD_EXCEL,
                )
                row["market_value"] = _parse_number(raw[3])
            else:
                row = _jpmorgan_non_stock_row(
                    raw,
                    asset_type,
                    etf_code,
                    source_url,
                    date_str,
                )
            if row:
                rows.append(row)

    rows = dedupe_rows(rows)
    if len([row for row in rows if row["asset_type"] == "stock"]) < 5:
        raise ValueError("JPMorgan stock rows not found")
    return rows


def scrape_jpmorgan_excel(etf_code: str, target_date: date) -> dict:
    etf_code = etf_code.upper()
    config = get_official_config(etf_code)
    source_url = config["url"]
    try:
        params = {**config["internal_ids"], "date": target_date.isoformat()}
        response = requests.get(
            source_url,
            params=params,
            headers={
                "User-Agent": "Mozilla/5.0",
                "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
            timeout=30,
        )
        response.raise_for_status()
        source_url = response.url
        rows = parse_jpmorgan_excel(
            response.content,
            etf_code,
            source_url,
            target_date,
        )
        return _build_result(rows, source_url, EXTRACTION_METHOD_EXCEL)
    except Exception as exc:
        return _failed_result(source_url, f"JPMorgan Excel failed: {exc}")


def _parse_float(value: str) -> float | None:
    cleaned = value.strip().replace(",", "").replace("%", "")
    if not cleaned or cleaned.upper() in {"-", "--", "N/A", "NA"}:
        return None
    return float(cleaned)


def _parse_number(value: str) -> int | float | None:
    cleaned = value.strip().replace(",", "")
    if not cleaned or cleaned.upper() in {"-", "--", "N/A", "NA"}:
        return None
    number = float(cleaned)
    return int(number) if number.is_integer() else number


def _parser_for_issuer(issuer: str):
    parsers = {"Fubon": parse_fubon, "Taishin": parse_taishin, "TWSE": parse_twse}
    try:
        return parsers[issuer]
    except KeyError as exc:
        raise ValueError(f"No static official parser for issuer: {issuer}") from exc


def _parse_official_logic(logic: str) -> dict:
    internal_ids = {}
    for part in logic.split(";"):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        internal_ids[key.strip()] = value.strip()
    return internal_ids


def _build_twse_url(etf_code: str) -> str:
    return TWSE_URL_TEMPLATE.format(code=etf_code.upper())


def _sum_weights(rows: list) -> float:
    return round(sum(row["weight_pct"] for row in rows if row.get("weight_pct") is not None), 2)


def _validate_official_rows(rows: list) -> tuple[bool, str]:
    return validate_snapshot_rows(rows)


def _official_weight_warning(total_weight: float) -> dict | None:
    if total_weight < OFFICIAL_WARNING_MIN_TOTAL_WEIGHT:
        reason = "total_weight_below_expected_range"
    elif total_weight > OFFICIAL_WARNING_MAX_TOTAL_WEIGHT:
        reason = "total_weight_above_expected_range"
    else:
        return None
    return {
        "reason": reason,
        "source_total_weight_all_rows": total_weight,
        "minimum_expected_weight": OFFICIAL_WARNING_MIN_TOTAL_WEIGHT,
        "maximum_expected_weight": OFFICIAL_WARNING_MAX_TOTAL_WEIGHT,
    }



def _failed_result(source_url: str, reason: str) -> dict:
    return {
        "ok": False,
        "reason": reason,
        "all_rows": [],
        "stock_rows": [],
        "non_stock_rows": [],
        "source_url": source_url,
        "source_type": SOURCE_TYPE,
        "total_weight_all_rows": 0.0,
        "total_weight_stock_rows": 0.0,
    }


def _build_result(all_rows: list, source_url: str, extraction_method: str) -> dict:
    ok, reason = _validate_official_rows(all_rows)
    stock_rows, non_stock_rows = split_rows(all_rows)
    total_weight_all_rows = _sum_weights(all_rows)
    result = {
        "ok": ok,
        "reason": reason,
        "all_rows": all_rows,
        "stock_rows": stock_rows,
        "non_stock_rows": non_stock_rows,
        "source_url": source_url,
        "source_type": SOURCE_TYPE,
        "total_weight_all_rows": total_weight_all_rows,
        "total_weight_stock_rows": _sum_weights(stock_rows),
    }
    if ok:
        warning = _official_weight_warning(total_weight_all_rows)
        if warning is not None:
            result["weight_warning"] = warning
    return result
